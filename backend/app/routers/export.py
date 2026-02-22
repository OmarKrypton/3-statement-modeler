from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from ..database import get_db
from .forecast import get_forecast_statements
from .statements import get_income_statement, get_balance_sheet, get_cash_flow

router = APIRouter(
    prefix="/api/v1/companies/{company_id}/export",
    tags=["Export"]
)

# Shared styles
bold_font = Font(bold=True)
header_font = Font(bold=True, size=12, color="FFFFFF")
title_font = Font(bold=True, size=16, color="333333")
right_align = Alignment(horizontal="right")
center_align = Alignment(horizontal="center", wrap_text=True, vertical="center")
bottom_border = Border(bottom=Side(style='thin', color="CCCCCC"))
all_border = Border(bottom=Side(style='thin', color="EEEEEE"), top=Side(style='thin', color="EEEEEE"))
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)' 
# We'll use a slightly cleaner version to ensure brackets are obvious
# Standard Excel Accounting format (with brackets for negatives and $ alignment)
accounting_format = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"_);_(@_)'

def add_title(sheet, title_text, col_span):
    sheet.append([title_text])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    cell = sheet.cell(row=1, column=1)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 25

def add_header(sheet, columns):
    sheet.append(columns)
    row = sheet.max_row
    for cell in sheet[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = bottom_border
        if cell.column > 1:
            cell.alignment = center_align

def format_row(sheet, row_idx, is_bold=False):
    for cell in sheet[row_idx]:
        if cell.column > 1:
            cell.number_format = accounting_format
        if is_bold:
            cell.font = bold_font

@router.get("/excel")
def export_forecast_excel(
    company_id: str, 
    scenario: str = "base", 
    include_is: bool = True,
    include_bs: bool = True,
    include_cf: bool = True,
    db: Session = Depends(get_db)
):
    # 1. Fetch data dictionary from the existing forecasting engine
    data = get_forecast_statements(company_id, scenario, db)
    
    actuals = data.get("actuals", {})
    projections = data.get("projections", [])
    if not projections:
        raise HTTPException(status_code=400, detail="No projection data found to export.")

    base_period = data["base_period"]
    
    # Header row: 'Metric', 'Actuals (Date)', 'Forecast (Date1)', 'Forecast (Date2)'...
    headers = ["Metric", f"Actuals\n{base_period}"]
    for p in projections:
        headers.append(f"Forecast\n{p['period']}")

    # 2. Setup Workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    total_cols = len(headers)
    
    # ── INCOME STATEMENT ──
    if include_is:
        ws_is = wb.create_sheet("Income Statement")
        add_title(ws_is, f"Automated 3-Statement Modeler - Income Statement ({scenario.capitalize()} Scenario)", total_cols)
        add_header(ws_is, headers)
        ws_is.freeze_panes = "B3"
        
        def r(actual_val, key, flip_sign=False):
            row = [actual_val]
            for p in projections:
                val = p[key] / 100.0  # Convert cents to dollars
                row.append(-val if flip_sign else val)
            return row

        ws_is.append(["Revenue"] + r(actuals.get("revenue_cents", 0) / 100.0, "revenue_cents"))
        ws_is.append(["COGS"] + r(-(actuals.get("revenue_cents", 0)*0 / 100.0), "cogs_cents", flip_sign=True)) # Just ensuring logic
        ws_is.append(["Gross Profit"] + r(0, "gross_profit_cents"))
        format_row(ws_is, ws_is.max_row, is_bold=True)
        
        ws_is.append(["Operating Expenses"] + r(-(actuals.get("expenses_cents", 0) / 100.0), "opex_cents", flip_sign=True))
        ws_is.append(["EBITDA"] + r(0, "ebitda_cents"))
        format_row(ws_is, ws_is.max_row, is_bold=True)
        
        ws_is.append(["D&A"] + r(0, "da_cents", flip_sign=True))
        ws_is.append(["EBIT"] + r(0, "ebit_cents"))
        format_row(ws_is, ws_is.max_row, is_bold=True)
        
        ws_is.append(["Tax"] + r(0, "tax_cents", flip_sign=True))
        ws_is.append(["Net Income"] + r(actuals.get("net_income_cents", 0) / 100.0, "net_income_cents"))
        format_row(ws_is, ws_is.max_row, is_bold=True)
    
    # ── BALANCE SHEET (Placeholder for now, but following the pattern) ──
    if include_bs:
        ws_bs = wb.create_sheet("Balance Sheet")
        add_title(ws_bs, f"Automated 3-Statement Modeler - Balance Sheet ({scenario.capitalize()} Scenario)", total_cols)
        add_header(ws_bs, headers)
        ws_bs.freeze_panes = "B3"
        # Since we don't have full BS projection metrics in the engine yet, we'll just add headers and a note
        ws_bs.append(["Total Assets", 0] + [0]*len(projections))
        ws_bs.append(["Total Liabilities", 0] + [0]*len(projections))
        ws_bs.append(["Total Equity", 0] + [0]*len(projections))
        format_row(ws_bs, ws_bs.max_row, is_bold=True)
    
    # ── CASH FLOW STATEMENT ──
    if include_cf:
        ws_cf = wb.create_sheet("Cash Flow")
        add_title(ws_cf, f"Automated 3-Statement Modeler - Cash Flow ({scenario.capitalize()} Scenario)", total_cols)
        add_header(ws_cf, headers)
        ws_cf.freeze_panes = "B3"
        
        # Internal helper for CF rows
        def r_cf(actual_val, key, flip_sign=False):
            row = [actual_val]
            for p in projections:
                val = p[key] / 100.0
                row.append(-val if flip_sign else val)
            return row
    
        ws_cf.append(["Net Income"] + r_cf(actuals.get("net_income_cents", 0) / 100.0, "net_income_cf_cents"))
        ws_cf.append(["D&A (Add-back)"] + r_cf(0, "da_cents"))
        ws_cf.append(["Δ Net Working Capital"] + r_cf(0, "delta_wc_cents"))
        ws_cf.append(["Cash from Operations"] + r_cf(0, "net_cash_from_operations_cents"))
        format_row(ws_cf, ws_cf.max_row, is_bold=True)
        
        ws_cf.append(["CapEx"] + r_cf(0, "capex_cents"))
        ws_cf.append(["Cash from Investing"] + r_cf(0, "net_cash_from_investing_cents"))
        format_row(ws_cf, ws_cf.max_row, is_bold=True)
        
        ws_cf.append(["Cash from Financing"] + r_cf(0, "net_cash_from_financing_cents"))
        format_row(ws_cf, ws_cf.max_row, is_bold=True)
        
        ws_cf.append(["Net Change in Cash"] + r_cf(0, "net_change_in_cash_cents"))
        ws_cf.append(["Beginning Cash"] + r_cf(actuals.get("cash_cents", 0) / 100.0, "beginning_cash_cents"))
        ws_cf.append(["Ending Cash"] + r_cf(actuals.get("cash_cents", 0) / 100.0, "ending_cash_cents"))
        format_row(ws_cf, ws_cf.max_row, is_bold=True)

    # Clean up column widths
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 45 # Even wider for "Automated 3-Statement Modeler..."
        for col_idx in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # 3. Stream back to client
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Construct descriptive filename
    mapping = {
        "IS": "Income_Statement",
        "BS": "Balance_Sheet",
        "CF": "Cash_Flow"
    }
    selected = [mapping[s] for s in ["IS", "BS", "CF"] if (include_is if s=="IS" else include_bs if s=="BS" else include_cf)]
    
    if not selected:
        filename = f"Forecast_{scenario.capitalize()}.xlsx"
    elif len(selected) == 3:
        filename = f"Full_Forecast_{scenario.capitalize()}.xlsx"
    else:
        filename = f"{'_'.join(selected)}_{scenario.capitalize()}.xlsx"

    headers_response = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response
    )

@router.get("/actuals/excel")
def export_actuals_excel(
    company_id: str, 
    periods: str,  # Comma-separated periods
    include_is: bool = True,
    include_bs: bool = True,
    include_cf: bool = True,
    db: Session = Depends(get_db)
):
    # Convert string periods to date objects
    try:
        period_list = [date.fromisoformat(p) for p in periods.split(",")]
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid period format. Expected ISO (YYYY-MM-DD)")
    
    # Fetch Data
    is_data = get_income_statement(company_id, period_list, db)
    bs_data = get_balance_sheet(company_id, period_list, db)
    cf_data = get_cash_flow(company_id, period_list, db)
    
    # Header row
    headers = ["Metric"] + period_list
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default
    total_cols = len(headers)
    
    if include_is:
        ws = wb.create_sheet("Income Statement")
        add_title(ws, "Financial Report - Income Statement (Actuals)", total_cols)
        add_header(ws, headers)
        ws.freeze_panes = "B3"
        
        # We need to pivot the data by metric
        metrics = [
            ("Total Revenues", "total_revenues_cents"),
            ("Total Operating Expenses", "total_expenses_cents"),
            ("Net Income", "net_income_cents")
        ]
        
        for label, key in metrics:
            row = [label]
            for p in periods.split(","):
                # Find the period data
                p_data = next((item for item in is_data if item["period"] == p), None)
                val = (p_data[key] / 100.0) if p_data else 0
                # Flip sign for expenses to show in brackets
                if key == "total_expenses_cents": val = -val
                row.append(val)
            ws.append(row)
            format_row(ws, ws.max_row, is_bold=(label == "Net Income"))

    if include_bs:
        ws = wb.create_sheet("Balance Sheet")
        add_title(ws, "Financial Report - Balance Sheet (Actuals)", total_cols)
        add_header(ws, headers)
        ws.freeze_panes = "B3"
        
        metrics = [
            ("Total Assets", "total_assets_cents"),
            ("Total Liabilities", "total_liabilities_cents"),
            ("Total Equity", "total_equity_cents")
        ]
        
        for label, key in metrics:
            row = [label]
            for p in periods.split(","):
                p_data = next((item for item in bs_data if item["period"] == p), None)
                val = (p_data[key] / 100.0) if p_data else 0
                # Flip sign for Liabilities and Equity (Credits) to show in brackets
                if key in ["total_liabilities_cents", "total_equity_cents"]: val = -val
                row.append(val)
            ws.append(row)
            format_row(ws, ws.max_row, is_bold=True)

    if include_cf:
        ws = wb.create_sheet("Cash Flow")
        add_title(ws, "Financial Report - Cash Flow (Actuals)", total_cols)
        add_header(ws, headers)
        ws.freeze_panes = "B3"
        
        metrics = [
            ("Net Income", "net_income_cents"),
            ("Depreciation & Non-Cash", "non_cash_adjustments_cents"),
            ("Changes in Working Capital", "operating_wc_delta_cents"),
            ("Net Cash from Ops", "net_cash_from_operations_cents"),
            ("Net Cash from Investing", "net_cash_from_investing_cents"),
            ("Net Cash from Financing", "net_cash_from_financing_cents"),
            ("End Balance Cash", "ending_cash_cents")
        ]
        
        for label, key in metrics:
            row = [label]
            for p in periods.split(","):
                p_data = next((item for item in cf_data if item["period"] == p), None)
                val = (p_data[key] / 100.0) if p_data else 0
                row.append(val)
            ws.append(row)
            format_row(ws, ws.max_row, is_bold=(label in ["Net Cash from Ops", "End Balance Cash"]))

    # Clean up
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 45
        for col_idx in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Construct descriptive filename
    mapping = {
        "IS": "Income_Statement",
        "BS": "Balance_Sheet",
        "CF": "Cash_Flow"
    }
    selected = [mapping[s] for s in ["IS", "BS", "CF"] if (include_is if s=="IS" else include_bs if s=="BS" else include_cf)]
    
    if not selected:
        filename = "Financial_Statements_Actuals.xlsx"
    elif len(selected) == 3:
        filename = "Full_Financial_Report_Actuals.xlsx"
    else:
        filename = f"{'_'.join(selected)}_Actuals.xlsx"

    headers_response = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response
    )
